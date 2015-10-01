from collections import defaultdict
from api.models.QueryParameters import QueryResult
from openpyxl import load_workbook


class DataMappingRepository:
    def __init__(self, uow):
        self.uow = uow
        self.table = uow.tables.data_mapping
        self.unknown_vendor_points_table = uow.tables.unknown_vendor_points

    def get_mapping_by_id(self, model_id):
        model = self.uow.run(self.table.get(model_id))
        return model

    def get_mapping_by_syrx_num(self, syrx_num):
        try:
            data_mapping_point = self.uow.run_list(self.table.get_all(syrx_num, index="syrx_num"))[0]
            return data_mapping_point
        except:
            return None

    def get_mappings_by_syrx_nums(self, syrx_nums):
        if not syrx_nums or len(syrx_nums) == 0:
            return []
        points = self.uow.run_list(self.table.get_all(*syrx_nums, index="syrx_num"))
        return points

    def get_mappings_for_johnson_site_id_by_fqr(self, site_id):
        mappings = self.uow.run_list(self.table.get_all(site_id, index="johnson_site_id"))
        mappings = dict((x["johnson_fqr"], x) for x in mappings)
        return mappings

    def get_mappings_for_johnson_imported_vendor_records(self):
        mappings = self.uow.run_list(self.uow.tables.imported_vendor_records
                                     .eq_join([self.uow.row_obj["johnson_site_id"], self.uow.row_obj["johnson_fqr"]],
                                              self.uow.tables.data_mapping, index="johnson")
                                     .without({"left": "date_added", "right": "id"}).zip()
                                     .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                     .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                     .eq_join("component_point_id", self.uow.tables.component_points)
                                     .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings

    def get_mappings_for_johnson_imported_vendor_records_cursor(self):
        mappings = self.uow.run(self.uow.tables.imported_vendor_records
                                .eq_join([self.uow.row_obj["johnson_site_id"], self.uow.row_obj["johnson_fqr"]],
                                         self.uow.tables.data_mapping, index="johnson")
                                .without({"left": "date_added", "right": "id"}).zip()
                                .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                .eq_join("component_point_id", self.uow.tables.component_points)
                                .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings

    def get_all_mappings_for_johnson(self):
        mappings = self.uow.run_list(self.uow.tables.data_mapping.get_all("johnson", index="source"))
        return mappings

    def get_mappings_for_johnson_site_id_fqr(self, keys):
        # keys should be of format [[site_id1, fqr1], [site_id2, fqr2], ...]
        if len(keys) == 0:
            return []
        mappings = self.uow.run(self.uow.tables.data_mapping
                                .get_all(*keys, index="johnson")
                                .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                .eq_join("component_point_id", self.uow.tables.component_points)
                                .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings

    def get_all_mappings_for_invensys(self):
        mappings = self.uow.run_list(self.uow.tables.data_mapping.get_all("invensys", index="source"))
        return mappings

    def get_mappings_for_invensys_site_equipment_point(self, keys):
        # keys should be of format [[site_name, equipment_name, point_name], [...], ...]
        if len(keys) == 0:
            return []
        mappings = self.uow.run(self.uow.tables.data_mapping
                                .get_all(*keys, index='invensys')
                                .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                .eq_join("component_point_id", self.uow.tables.component_points)
                                .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings

    def get_mappings_for_siemens_meter_name(self, keys):
        # keys should be of format [[siemens_meter_name], [siemens_meter_name], ...]
        if len(keys) == 0:
            return []
        mappings = self.uow.run(self.uow.tables.data_mapping
                                .get_all(*keys, index="siemens")
                                .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                .eq_join("component_point_id", self.uow.tables.component_points)
                                .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings

    def get_all_mappings_for_siemens(self):
        mappings = self.uow.run_list(self.uow.tables.data_mapping.get_all("siemens", index="source"))
        return mappings

    def get_all_mappings_for_fieldserver(self):
        mappings = self.uow.run_list(self.uow.tables.data_mapping.get_all("fieldserver", index="source"))
        return mappings

    def get_mappings_for_fieldserver_site_id_offset(self, keys):
        # keys should be of format [[site_id1, offset1], [site_id2, offset2], ...]
        if len(keys) == 0:
            return []
        mappings = self.uow.run(self.uow.tables.data_mapping
                                .get_all(*keys, index="fieldserver")
                                .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                                .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                                .eq_join("component_point_id", self.uow.tables.component_points)
                                .without({"right": ["id", "component_id", "description"]}).zip())

        return mappings



    def get_mappings_for_johnson_imported_vendor_records_count(self):
        count = self.uow.run(self.uow.tables.imported_vendor_records
                             .eq_join([self.uow.row_obj["johnson_site_id"], self.uow.row_obj["johnson_fqr"]],
                                      self.uow.tables.data_mapping, index="johnson")
                             .without({"left": "date_added", "right": "id"}).zip()
                             .eq_join("syrx_num", self.uow.tables.equipment_points, index="syrx_num")
                             .without({"right": ["equipment_id", "id", "syrx_num"]}).zip()
                             .eq_join("component_point_id", self.uow.tables.component_points)
                             .without({"right": ["id", "component_id", "description"]}).zip().count())

        return count

    def get_unknown_vendor_ponts_for_johnson_site_id_by_fqr(self, site_id):
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all(site_id, index="johnson_site_id"))
        vendor_points = dict((x["johnson_fqr"], x) for x in vendor_points)
        return vendor_points

    def get_unknown_vendor_points_for_johnson_site_id_fqr(self, keys):
        if len(keys) == 0:
            return []
        # keys should be of format [[site_id1, fqr1], [site_id2, fqr2], ...]
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all(*keys, index="johnson"))
        return vendor_points

    def get_unknown_vendor_points_for_invensys_site_equipment_point(self, keys):
        if len(keys) == 0:
            return []
        # keys should be of format [[site_name, equipment_name, point_name], [site_name, equipment_name, point_name], ...]
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all(*keys, index="invensys"))
        return vendor_points


    def get_unknown_vendor_points_for_siemens_meter_name(self, keys):
        if len(keys) == 0:
            return []
        # keys should be of format [[siemens_meter_name], [siemens_meter_name],, ...]
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all(*keys, index="siemens"))
        return vendor_points

    def get_unknown_vendor_points_for_fieldserver_site_id_offset(self, keys):
        if len(keys) == 0:
            return []
        # keys should be of format [[site_id1, offset1], [site_id2, offset2], ...]
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all(*keys, index="fieldserver"))
        return vendor_points

    def get_unknown_vendor_points_for_johnson(self):
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all("johnson", index="source"))
        return vendor_points

    def get_unknown_vendor_points_for_fieldserver(self):
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all("fieldserver", index="source"))
        return vendor_points

    def get_unknown_vendor_points_for_invensys(self):
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all("invensys", index="source"))
        return vendor_points

    def get_unknown_vendor_points_for_siemens(self):
        vendor_points = self.uow.run_list(self.unknown_vendor_points_table.get_all("siemens", index="source"))
        return vendor_points

    def apply_query_parameters(self, query_parameters, source):
        table = self.table.get_all(source, index="source")
        return self.uow.apply_query_parameters(table, query_parameters)

    def unknown_apply_query_parameters(self, query_parameters, source):
        table = self.unknown_vendor_points_table.get_all(source, index="source")
        return self.uow.apply_query_parameters(table, query_parameters)

    def unknown_johnson_apply_query_parameters(self, query_parameters):
        table = self.unknown_vendor_points_table.order_by(index="johnson")
        q = table


        cursor = self.uow.run(q.skip(query_parameters.skip).limit(query_parameters.take))
        data = list(cursor)
        total = self.uow.run(table.count())

        query_result = QueryResult(data, total)
        return query_result

    def unknown_fieldserver_apply_query_parameters(self, query_parameters):
        table = self.unknown_vendor_points_table.order_by(index="fieldserver")
        q = table


        cursor = self.uow.run(q.skip(query_parameters.skip).limit(query_parameters.take))
        data = list(cursor)
        total = self.uow.run(table.count())

        query_result = QueryResult(data, total)
        return query_result

    def unknown_invensys_apply_query_parameters(self, query_parameters):
        table = self.unknown_vendor_points_table.order_by(index="invensys")
        q = table

        cursor = self.uow.run(q.skip(query_parameters.skip).limit(query_parameters.take))
        data = list(cursor)
        total = self.uow.run(table.count())

        query_result = QueryResult(data, total)
        return query_result

    def unknown_siemens_apply_query_parameters(self, query_parameters):
        table = self.unknown_vendor_points_table.order_by(index="siemens")
        q = table

        cursor = self.uow.run(q.skip(query_parameters.skip).limit(query_parameters.take))
        data = list(cursor)
        total = self.uow.run(table.count())

        query_result = QueryResult(data, total)
        return query_result

    def insert_mapping(self, mapping, source):
        try:
            del mapping["id"]
        except KeyError:
            pass

        equipment_point = self.uow.equipment.get_equipment_point_by_syrx_num(mapping["syrx_num"])
        if equipment_point is None:
            raise Exception("Syrx num does not exist")
        mapping_point = self.get_mapping_by_syrx_num(mapping["syrx_num"])
        if mapping_point is not None:
            raise Exception("Syrx num already used")

        mapping["source"] = source

        result = self.uow.run(self.table.insert(mapping))
        mapping["id"] = result["generated_keys"][0]
        return mapping

    def update_mapping(self, id, mapping, source):
        equipment_point = self.uow.equipment.get_equipment_point_by_syrx_num(mapping["syrx_num"])
        if equipment_point is None:
            raise Exception("Syrx num does not exist")
        mapping_point = self.get_mapping_by_syrx_num(mapping["syrx_num"])
        if mapping_point is not None and mapping_point["id"] != id:
            raise Exception("Syrx num already used")
        if mapping_point is not None and mapping_point["source"] != source:
            raise Exception("Syrx num is invalid")

        self.uow.run(self.table.get(id).update(mapping))

    def delete_mapping(self, model_id):
        self.uow.run(self.table.get(model_id).delete())

    def insert_unknown_vendor_points(self, unknown_vendor_points):
        self.uow.run(self.unknown_vendor_points_table.insert(unknown_vendor_points))

    def remove_unknown_johnson_vendor_point(self, site_id, fqr):
        self.uow.run(self.unknown_vendor_points_table.get_all([site_id, fqr], index="johnson").delete())

    def remove_unknown_fieldserver_vendor_point(self, site_id, fqr):
        self.uow.run(self.unknown_vendor_points_table.get_all([site_id, fqr], index="fieldserver").delete())

    def remove_unknown_invensys_vendor_point(self, site_name, equipment_name, point_name):
        self.uow.run(self.unknown_vendor_points_table.get_all([site_name, equipment_name, point_name], index="invensys").delete())

    def remove_unknown_siemens_vendor_point(self, siemens_meter_name):
        self.uow.run(self.unknown_vendor_points_table.get_all([siemens_meter_name], index="siemens").delete())

    def insert_imported_equipment_point_records(self, records):
        self.uow.run(self.uow.tables.imported_equipment_point_records.insert(records))

    def get_imported_equipment_point_records(self):
        return self.uow.run_list(self.uow.tables.imported_equipment_point_records)

    def delete_imported_equipment_point_records(self, record_ids):
        if len(record_ids) > 0:
            self.uow.run(self.uow.tables.imported_equipment_point_records.get_all(*record_ids).delete())

    def insert_imported_vendor_records(self, records):
        self.uow.run(self.uow.tables.imported_vendor_records.insert(records))

    def get_imported_vendor_records(self, source, skip=None, limit=None):
        q = self.uow.tables.imported_vendor_records.get_all(source, index="source")
        if skip is not None:
            q = q.skip(skip)
        if limit is not None:
            q = q.limit(limit)
        return self.uow.run_list(q)

    def get_imported_vendor_records_cursor(self, source):
        return self.uow.run(self.uow.tables.imported_vendor_records.get_all(source, index="source"))

    def get_imported_vendor_records_count(self, source):
        return self.uow.run(self.uow.tables.imported_vendor_records.get_all(source, index="source").count())

    def delete_imported_vendor_records(self, record_ids):
        if len(record_ids) > 0:
            self.uow.run(self.uow.tables.imported_vendor_records.get_all(*record_ids).delete())


    class UploadResults:
        def __init__(self):
            self.added_mappings = []
            self.johnson_point_already_mapped = []
            self.syrx_num_already_mapped = []
            self.syrx_num_doesnt_exist = []


    def upload_johnson_mappings(self, excel_filename):
        wb = load_workbook(filename=excel_filename, use_iterators=True)
        ws = wb.worksheets[0]

        # map out rows to the format of data_mapping
        rows = map(lambda x: {
                "source": "johnson",
                "johnson_site_id": x[0].value,
                "johnson_fqr": x[1].value,
                "syrx_num": x[2].value
        }, ws.iter_rows())

        # remove header row
        rows = rows[1:]

        # build list of site_id/fqr pairs from the excel file to easily query the database index
        site_id_fqr_pairs = map(lambda x: [x["johnson_site_id"], x["johnson_fqr"]], rows)

        # build a two-level dictionary of existing johnson point mappings based on the points included in the excel file
        # this dictionary is of form (site_id) => (fqr) => (mapping)
        # to determine if a mapping from the excel file is using a johnson point that already exists, we check
        # this dictionary. if it contains a key that matches the site_id, we check the returned dictionary. if that
        # dictionary contains a key that matches the fqr, then we know there is an existing mapping.
        johnson_point_existing_mappings = self.uow.run_list(self.table.get_all(*site_id_fqr_pairs, index="johnson"))
        johnson_point_existing_mappings_dict = defaultdict(dict)

        for x in johnson_point_existing_mappings:
            johnson_point_existing_mappings_dict[x["johnson_site_id"]][x["johnson_fqr"]] = x
        johnson_point_existing_mappings = johnson_point_existing_mappings_dict

        # build a dictionary of syrx nums from the excel file to easily query the database index
        syrx_nums = map(lambda x: x["syrx_num"], rows)

        # build a dictionary of existing johnson point mappings based on the points included in the excel file
        # this dictionary is of form (syrx_num) => (mapping)
        # to determine if a mapping from the excel file is using a syrx num that already has a vendor point mapping,
        # we check if this dictionary has a key that matches the syrx num.
        syrx_num_point_existing_mappings = self.uow.run_list(self.table.get_all(*syrx_nums, index="syrx_num"))
        syrx_num_point_existing_mappings = dict((x["syrx_num"], x) for x in syrx_num_point_existing_mappings)


        # build a dictionary of equipment points from the database based on syrx nums included in the excel file
        # this dictionary is of form (syrx_num) => (equipment_point)
        # to determine if a mapping from the excel file is using an existing equipment point, we check if
        # the dictionary contains a key that matches the syrx num
        existing_equipment_points = self.uow.run_list(self.uow.tables.equipment_points.get_all(*syrx_nums, index="syrx_num"))
        existing_equipment_points = dict((x["syrx_num"], x) for x in existing_equipment_points)

        mappings_to_add = []
        johnson_point_already_mapped = []
        syrx_num_already_mapped = []
        syrx_num_doesnt_exist = []

        for row in rows:
            #good_mapping = True
            if row["syrx_num"] not in existing_equipment_points:
                syrx_num_doesnt_exist.append(row)
                continue
                #good_mapping = False
            if row["johnson_site_id"] in johnson_point_existing_mappings \
                    and row["johnson_fqr"] in johnson_point_existing_mappings[row["johnson_site_id"]]:
                johnson_point_already_mapped.append(row)
                continue
                #good_mapping = False
            if row["syrx_num"] in syrx_num_point_existing_mappings:
                syrx_num_already_mapped.append(row)
                continue
                #good_mapping = False
            #if good_mapping:
            mappings_to_add.append(row)

        self.uow.run(self.table.insert(mappings_to_add))

        # delete the successful mappings from the unknown mappings table
        successful_site_id_fqr_pairs = [[x["johnson_site_id"], x["johnson_fqr"]] for x in mappings_to_add]
        if len(successful_site_id_fqr_pairs) > 0:
            self.uow.run(self.unknown_vendor_points_table.get_all(*successful_site_id_fqr_pairs, index="johnson").delete())

        results = self.UploadResults()
        results.added_mappings = mappings_to_add
        results.johnson_point_already_mapped = johnson_point_already_mapped
        results.syrx_num_already_mapped = syrx_num_already_mapped
        results.syrx_num_doesnt_exist = syrx_num_doesnt_exist

        return results

    def make_siemens_point_global(self, siemens_meter_name):
        self.uow.run(self.unknown_vendor_points_table.get_all([siemens_meter_name], index="siemens")
                     .update({"global": True}))

    def make_johnson_point_global(self, johnson_site_id, johnson_fqr):
        self.uow.run(self.unknown_vendor_points_table.get_all([johnson_site_id, johnson_fqr], index='johnson')
                     .update({'global': True}))

    def make_fieldserver_point_global(self, fieldserver_site_id, fieldserver_offset):
        self.uow.run(self.unknown_vendor_points_table.get_all([fieldserver_site_id, fieldserver_offset],
                                                              index='fieldserver')
                     .update({'global': True}))

    def make_invensys_point_global(self, invensys_site_name, invensys_equipment_name, invensys_point_name):
        self.uow.run(self.unknown_vendor_points_table.get_all([invensys_site_name, invensys_equipment_name,
                                                               invensys_point_name], index='invensys')
                     .update({'global': True}))

    def remove_global_property(self, vendor_point):
        if 'siemens_meter_name' in vendor_point:
            self.uow.run(self.unknown_vendor_points_table.get_all('siemens', index='source')
                         .filter({'siemens_meter_name': vendor_point['siemens_meter_name']})
                         .replace(self.uow.row_obj.without('global')))
        elif 'johnson_site_id' in vendor_point and 'johnson_fqr' in vendor_point:
            self.uow.run(self.unknown_vendor_points_table.get_all('johnson', index='source')
                         .filter({'johnson_site_id': vendor_point['johnson_site_id'],
                                  'johnson_fqr': vendor_point['johnson_fqr']})
                         .replace(self.uow.row_obj.without('global')))
        elif 'fieldserver_site_id' in vendor_point and 'fieldserver_offset' in vendor_point:
            self.uow.run(self.unknown_vendor_points_table.get_all('fieldserver', index='source')
                         .filter({'fieldserver_site_id': vendor_point['fieldserver_site_id'],
                                  'fieldserver_offset': vendor_point['fieldserver_offset']})
                         .replace(self.uow.row_obj.without('global')))
        elif 'invensys_site_name' in vendor_point and 'invensys_equipment_name' in vendor_point and 'invensys_point_name' in vendor_point:
            self.uow.run(self.unknown_vendor_points_table.get_all('invensys', index='source')
                         .filter({'invensys_site_name': vendor_point['invensys_site_name'],
                                  'invensys_equipment_name': vendor_point['invensys_equipment_name'],
                                  'invensys_point_name': vendor_point['invensys_point_name']})
                         .replace(self.uow.row_obj.without('global')))

    def get_all_global_vendor_points(self):
        return self.uow.run_list(self.table.filter({'global': True}))