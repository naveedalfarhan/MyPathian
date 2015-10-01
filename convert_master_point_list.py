from openpyxl import load_workbook

energy_points_dict = dict()
calculated_points_dict = dict()
position_points_dict = dict()
binary_points_dict = dict()
numeric_points_dict = dict()

def load_energy_points_dict(wb):
    energy_points_ws = wb.get_sheet_by_name("Energy Points")

    for row_num, row in enumerate(energy_points_ws.iter_rows()):
        if row_num < 2:
            continue

        energy_point = {
            "num": str(row[1].value).strip(),
            "code": str(row[1].value).strip(),
            "description": str(row[0].value).strip(),
            "units": str(row[2].value).strip()
        }

        if energy_point["num"] is None or len(energy_point["num"]) == 0:
            continue

        energy_points_dict[energy_point["num"]] = energy_point

def load_calculated_points_dict(wb):
    ws = wb.get_sheet_by_name("Calculation Points")

    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < 3:
            continue

        point = {
            "num": str(row[1].value).strip(),
            "code": str(row[1].value).strip(),
            "description": str(row[0].value).strip(),
            "formula": str(row[2].value).strip(),
            "units": ""
        }

        if point["num"] is None or len(point["num"]) == 0:
            continue

        calculated_points_dict[point["num"]] = point

def load_position_points_dict(wb):
    ws = wb.get_sheet_by_name("Position Points")

    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < 3:
            continue

        point = {
            "num": str(row[1].value).strip(),
            "code": str(row[1].value).strip(),
            "description": str(row[0].value).strip()
        }

        if point["num"] is None or len(point["num"]) == 0:
            continue

        position_points_dict[point["num"]] = point

def load_binary_points_dict(wb):
    ws = wb.get_sheet_by_name("Binary Points")

    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < 2:
            continue

        point = {
            "num": str(row[1].value).strip(),
            "code": str(row[1].value).strip(),
            "description": str(row[0].value).strip()
        }

        if point["num"] is None or len(point["num"]) == 0:
            continue

        binary_points_dict[point["num"]] = point

def load_numeric_points_dict(wb):
    ws = wb.get_sheet_by_name("Numeric Points")

    for row_num, row in enumerate(ws.iter_rows()):
        if row_num < 2:
            continue

        point = {
            "num": str(row[1].value).strip(),
            "code": str(row[1].value).strip(),
            "description": str(row[0].value).strip(),
            "units": ""
        }

        if point["num"] is None or len(point["num"]) == 0:
            continue

        numeric_points_dict[point["num"]] = point

def write_energy_points_csv_file(wb):
    components_ws = wb.get_sheet_by_name("Master Comp List Rev 1")
    with open("points_energy_points.csv", "w") as f:
        f.write("Component num,Point num,Code,Description,Units\n")
        for row_num, row in enumerate(components_ws.iter_rows()):
            if row_num < 1:
                continue

            if row[0].value is None:
                continue

            component_num = str(row[0].value).strip()
            energy_point_num = str(row[4].value).strip()

            if component_num is None or len(component_num) == 0:
                continue

            try:
                energy_point = energy_points_dict[energy_point_num]
                s = ",".join([component_num, energy_point["num"], energy_point["num"],
                              '"' + energy_point["description"] + '"', energy_point["units"]])
                f.write(s + "\n")
            except:
                pass

def write_calculated_points_csv_file(wb):
    components_ws = wb.get_sheet_by_name("Master Comp List Rev 1")
    with open("points_calculated_points.csv", "w") as f:
        f.write("Component num,Point num,Code,Description,Units,Formula\n")
        for row_num, row in enumerate(components_ws.iter_rows()):
            if row_num < 1:
                continue

            if row[0].value is None:
                continue

            component_num = str(row[0].value).strip()
            point_num = str(row[10].value).strip()

            if component_num is None or len(component_num) == 0:
                continue

            try:
                point = calculated_points_dict[point_num]
                s = ",".join([component_num, point["num"], point["num"],
                              '"' + point["description"] + '"', point["units"], '"' + point["formula"] + '"'])
                f.write(s + "\n")
            except:
                pass

def write_position_points_csv_file(wb):
    components_ws = wb.get_sheet_by_name("Master Comp List Rev 1")
    with open("points_position_points.csv", "w") as f:
        f.write("Component num,Point num,Code,Description\n")
        for row_num, row in enumerate(components_ws.iter_rows()):
            if row_num < 1:
                continue

            if row[0].value is None:
                continue

            component_num = str(row[0].value).strip()
            point_num = str(row[6].value).strip()

            if component_num is None or len(component_num) == 0:
                continue

            try:
                point = position_points_dict[point_num]
                s = ",".join([component_num, point["num"], point["num"],
                              '"' + point["description"] + '"'])
                f.write(s + "\n")
            except:
                pass

def write_binary_points_csv_file(wb):
    components_ws = wb.get_sheet_by_name("Master Comp List Rev 1")
    with open("points_binary_points.csv", "w") as f:
        f.write("Component num,Point num,Code,Description\n")
        for row_num, row in enumerate(components_ws.iter_rows()):
            if row_num < 1:
                continue

            if row[0].value is None:
                continue

            component_num = str(row[0].value).strip()
            point_num = str(row[8].value).strip()

            if component_num is None or len(component_num) == 0:
                continue

            try:
                point = binary_points_dict[point_num]
                s = ",".join([component_num, point["num"], point["num"],
                              '"' + point["description"] + '"'])
                f.write(s + "\n")
            except:
                pass

def write_numeric_points_csv_file(wb):
    components_ws = wb.get_sheet_by_name("Master Comp List Rev 1")
    with open("points_numeric_points.csv", "w") as f:
        f.write("Component num,Point num,Code,Description,Units\n")
        for row_num, row in enumerate(components_ws.iter_rows()):
            if row_num < 1:
                continue

            if row[0].value is None:
                continue

            component_num = str(row[0].value).strip()

            if component_num is None or len(component_num) == 0:
                continue

            for col_num in [12, 14]:
                point_num = str(row[col_num].value).strip()

                try:
                    point = numeric_points_dict[point_num]
                    s = ",".join([component_num, point["num"], point["num"],
                                  '"' + point["description"] + '"', point["units"]])
                    f.write(s + "\n")
                except:
                    pass


with open("E:\\pathian-uploads\\Master Point List.xlsx", "rb") as f:
    wb = load_workbook(filename=f, use_iterators=True)
    #load_energy_points_dict(wb)
    #load_calculated_points_dict(wb)
    #load_position_points_dict(wb)
    #load_binary_points_dict(wb)
    load_numeric_points_dict(wb)
    #write_energy_points_csv_file(wb)
    #write_calculated_points_csv_file(wb)
    #write_position_points_csv_file(wb)
    #write_binary_points_csv_file(wb)
    write_numeric_points_csv_file(wb)



