import pyodbc

from db.uow import UoW
from generate_flat_tables import FlatTableGenerator
from openpyxl import Workbook, load_workbook


__author__ = 'badams'


class CodeExporter():
    def __init__(self):
        self.sql_connection_string = "DRIVER={SQL Server};SERVER=.\SQLEXPRESS;DATABASE=PathianDBNet;integrated security=True"
        self.sql_connection = None
        self.workbook = None
        self.cursor = None

    def export_to_excel(self):
        # connect to sql
        self.connect_to_sql_db()

        # set up workbook
        self.create_workbook()

        # process all codes
        self.process_naics_codes()
        self.process_sic_codes()

        # save the workbook
        self.workbook.save(filename='naics_and_sic_codes.xlsx')

    def connect_to_sql_db(self):
        self.sql_connection = pyodbc.connect(self.sql_connection_string)

    def process_naics_codes(self):
        # add worksheet to workbook
        self.add_worksheet('naics')
        print "Adding NAICS worksheet."

        # get the proper worksheet
        ws = self.workbook.get_sheet_by_name('naics')

        # get the cursor
        self.get_naics_cursor()

        # add codes to worksheet
        self.add_codes_to_worksheet(ws)
        print "NAICS codes added to worksheet."

    def process_sic_codes(self):
        # add worksheet to workbook
        self.add_worksheet('sic')
        print "Adding SIC worksheet."

        # get the proper worksheet
        ws = self.workbook.get_sheet_by_name('sic')

        # get the cursor
        self.get_sic_cursor()

        # add codes to worksheet
        self.add_codes_to_worksheet(ws)
        print "SIC codes added to worksheet."

    def create_workbook(self):
        self.workbook = Workbook()
        remove_sheet = self.workbook.get_active_sheet()
        self.workbook.remove_sheet(remove_sheet)
        print "Workbook created."

    def add_worksheet(self, worksheet_name):
        ws = self.workbook.create_sheet()
        ws.title = worksheet_name

    def get_naics_cursor(self):
        self.cursor = self.sql_connection.cursor()
        self.cursor.execute("SELECT Code, Description, ParentCode FROM NaicsCodes")

    def add_codes_to_worksheet(self, ws):
        row_index = 1
        for row in self.cursor:
            ws.cell('A%d' % row_index).value = row.Code
            ws.cell('B%d' % row_index).value = row.Description
            ws.cell('C%d' % row_index).value = row.ParentCode
            ws.cell('D%d' % row_index).value = row.Code + ' - ' + row.Description
            ws.cell('E%d' % row_index).value = int(row.Code)
            row_index += 1

    def get_sic_cursor(self):
        self.cursor = self.sql_connection.cursor()
        self.cursor.execute("SELECT Code, Description, ParentCode FROM SicCodes")


class CodeImporter():
    def __init__(self):
        self.uow = UoW(False)
        self.filename = "naics_and_sic_codes.xlsx"
        self.workbook = None

    def run(self):
        # load the workbook
        self.get_workbook()

        # insert NAICS codes
        self.import_naics_codes()

        # insert SIC codes
        self.import_sic_codes()

    def get_workbook(self):
        self.workbook = load_workbook(self.filename, use_iterators=True)
        print "Workbook loaded."

    def import_naics_codes(self):
        # delete the old naics codes
        self.uow.naics.delete_all()

        # add the naics codes from the xlsx sheet to a list for a bulk insert
        codes = self.get_codes_from_sheet('naics')

        # insert the NAICS codes
        self.uow.naics.insert_batch(codes)

        print "NAICS codes imported."

    def import_sic_codes(self):
        # delete the old sic codes
        self.uow.sic.delete_all()

        # add the sic codes from the xlsx sheet to a list for a bulk insert
        codes = self.get_codes_from_sheet('sic')

        # insert the SIC codes
        self.uow.sic.insert_batch(codes)

        print "SIC codes imported."

    def get_codes_from_sheet(self, worksheet_name):
        ws = self.workbook.get_sheet_by_name(worksheet_name)
        codes = []
        for row in ws.iter_rows():
            codes.append(self.convert_row_to_dict(row))
        return codes

    @staticmethod
    def convert_row_to_dict(row):
        return {"code": row[0].value, "description": row[1].value, "parent_code": row[2].value,
                "name": row[3].value, "numeric_code": int(row[4].value)}


if __name__ == "__main__":
    importer = CodeImporter()
    importer.run()

    # generate flat tables
    generator = FlatTableGenerator()
    generator.generate_flat_groups()
    generator.generate_flat_naics_codes()
    generator.generate_flat_sic_codes()