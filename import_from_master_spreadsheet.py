import logging
import threading
import uuid
import math
from db.uow import UoW
from openpyxl import load_workbook


class MasterSpreadsheetImporter(threading.Thread):

    dry_threads = dict()
    wet_threads = dict()

    @classmethod
    def create(cls, file_path, importer_id=None, dry=True):
        t = MasterSpreadsheetImporter(file_path)
        t.dry = dry
        if dry:
            cls.dry_threads[importer_id] = t
        else:
            cls.wet_threads[importer_id] = t

        if importer_id is None:
            importer_id = uuid.uuid4()
        return importer_id

    @classmethod
    def start_dry(cls, importer_id):
        t = cls.dry_threads[importer_id]
        t.start()

    @classmethod
    def start_wet(cls, importer_id):
        t = cls.wet_threads[importer_id]
        t.start()

    @classmethod
    def get_dry_thread(cls, importer_id):
        return cls.dry_threads[importer_id]

    @classmethod
    def get_wet_thread(cls, importer_id):
        return cls.wet_threads[importer_id]

    def __init__(self, file_path):
        super(MasterSpreadsheetImporter, self).__init__()

        self.uow = UoW(None)
        self.wb = None
        self.file_path = file_path

        self.components_by_num = dict()
        self.points_by_num = dict()
        self.paragraphs_by_num = dict()

        self.results = {"components": [], "energy_points": [], "variable_points": [], "calculated_points": [], "position_points": [],
                        "numeric_points": [], "binary_points": []}

        self.current_progress_point = 0
        self.total_progress_points = 1

        self.running = False
        self.finished = False
        self.error = False

        self.dry = True

    def count_progress_points(self):
        self.total_progress_points = 0
        self.total_progress_points += self.count_rows_for_spreadsheet("Components") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Energy Points") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Variable Points") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Calculated Points") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Position Points") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Numeric Points") * 2
        self.total_progress_points += self.count_rows_for_spreadsheet("Binary Points") * 2

    def count_rows_for_spreadsheet(self, sheet_name):
        ws = self.wb.get_sheet_by_name(sheet_name)
        if ws is None:
            return 0
        return max(ws.max_row - 1, 0)

    def load_components_from_db(self):
        components = self.uow.components.get_all()
        self.components_by_num = dict((c["num"], c) for c in components)

    def load_points_from_db(self):
        points = self.uow.component_points.get_all()
        self.points_by_num = dict((p["component_point_num"], p) for p in points)

    def run(self):
        try:
            self.running = True

            self.wb = load_workbook(filename=self.file_path, use_iterators=True)
            self.count_progress_points()
            self.load_components_from_db()
            self.import_components()
            self.load_points_from_db()
            self.insert_energy_points()
            self.insert_variable_points()
            self.insert_calculated_points()
            self.insert_position_points()
            self.insert_numeric_points()
            self.insert_binary_points()
            """
            self.load_paragraphs_from_db()
            self.insert_paragraphs("AR")
            self.insert_paragraphs("CR")
            self.insert_paragraphs("CS")
            self.insert_paragraphs("DR")
            self.insert_paragraphs("FT")
            self.insert_paragraphs("LC")
            self.insert_paragraphs("MR")
            self.insert_paragraphs("PR")
            self.insert_paragraphs("RR")
            """
        except:
            logging.exception("An error occurred importing from master spreadsheet")
            self.error = True
        finally:
            self.running = False
            self.finished = True

    def import_components(self):
        ws = self.wb.get_sheet_by_name("Components")
        if ws is None:
            return

        rows = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue
            self.current_progress_point += 1

            rows.append({
                "component_num": str(row[0].value).strip(),
                "description": unicode(row[1].value).strip(),
                "component_id": str(uuid.uuid4())
            })

        # by sorting the list, we ensure that parents will always
        # come before their children
        rows = sorted(rows, key=lambda x: x["component_num"])

        components_to_insert = []

        for row in rows:
            self.current_progress_point += 1
            component_num = row["component_num"]
            description = row["description"]
            component_id = row["component_id"]

            if component_num in self.components_by_num:
                old_description = self.components_by_num[component_num]["description"]
                if description != old_description:
                    self.components_by_num[component_num]["description"] = description
                    components_to_insert.append(self.components_by_num[component_num])
                    self.results["components"].append({
                        "component_num": component_num,
                        "exists": True,
                        "action": True,
                        "error": False,
                        "result": "Component exists, changing description from '" + old_description + "' to '" + description + "'"
                    })
                else:
                    self.results["components"].append({
                        "component_num": component_num,
                        "exists": True,
                        "action": False,
                        "error": False,
                        "result": "Component exists, will be skipped"
                    })
                continue

            parent_num = component_num.strip()

            while len(parent_num) > 0 and parent_num[len(parent_num) - 1] not in [".", "-", " "]:
                parent_num = parent_num[0:len(parent_num)-1]

            parent_num = parent_num[0:len(parent_num)-1]

            if len(parent_num) == 5:
                # parent is level 2 which means it needs a level 2 multiple of 10
                parent_num = parent_num[0:4] + "0"

            try:
                parent = self.components_by_num[parent_num]
                parent_id = parent["id"]
            except:
                parent = None
                parent_id = None

            component = {
                "id": component_id,
                "description": description,
                "num": component_num,
                "mapping_child_ids": [],
                "mapping_parent_ids": [],
                "mapping_root": False,
                "protected": False,
                "structure_child_ids": [],
                "structure_parent_id": parent_id
            }

            if self.dry:
                component["from_spreadsheet"] = True

            components_to_insert.append(component)

            if parent is not None:
                parent["structure_child_ids"].append(component_id)
                self.results["components"].append({
                    "component_num": component_num,
                    "exists": False,
                    "action": True,
                    "error": False,
                    "result": "Will be added"
                })
            else:
                self.results["components"].append({
                    "component_num": component_num,
                    "exists": False,
                    "action": True,
                    "error": True,
                    "result": "Parent not found, component will be added at root level"
                })

            self.components_by_num[component_num] = component

        if not self.dry:
            batch = []
            for x in components_to_insert:
                batch.append(x)
                if len(batch) == 200:
                    self.uow.components.insert_bulk(batch)
                    batch = []
            if len(batch) > 0:
                self.uow.components.insert_bulk(batch)

    def insert_energy_points(self):
        ws = self.wb.get_sheet_by_name("Energy Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["energy_points"], "EP")

    def insert_variable_points(self):
        ws = self.wb.get_sheet_by_name("Variable Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["variable_points"], "VP")

    def insert_calculated_points(self):
        ws = self.wb.get_sheet_by_name("Calculated Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["calculated_points"], "CP")

    def insert_position_points(self):
        ws = self.wb.get_sheet_by_name("Position Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["position_points"], "PP")

    def insert_numeric_points(self):
        ws = self.wb.get_sheet_by_name("Numeric Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["numeric_points"], "NP")

    def insert_binary_points(self):
        ws = self.wb.get_sheet_by_name("Binary Points")
        if ws is None:
            return
        self.insert_points(ws, self.results["binary_points"], "BP")

    def insert_points(self, ws, results_array, point_type):
        rows = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue
            self.current_progress_point += 1

            point = self.parse_spreadsheet_row(row, point_type)
            if point:
                rows.append(point)

        points_to_insert = []

        for row in rows:
            self.current_progress_point += 1
            component_num = row["component_num"]
            point_num = row["point_num"]
            point_description = row["point_description"]

            if component_num not in self.components_by_num:
                results_array.append({
                    "point_num": point_num,
                    "exists": False,
                    "action": False,
                    "error": True,
                    "result": "Component not found, point will not be added"
                })
                continue

            component = self.components_by_num[component_num]

            if point_num in self.points_by_num:
                old_description = self.points_by_num[point_num]["description"]
                if point_description != old_description:
                    results_array.append({
                        "point_num": point_num,
                        "exists": True,
                        "action": True,
                        "error": False,
                        "result": "Point exists, changing description from '" + old_description + "' to '" +
                                  point_description + "'"
                    })
                else:
                    results_array.append({
                        "point_num": point_num,
                        "exists": True,
                        "action": False,
                        "error": False,
                        "result": "Point exists, will be skipped"
                    })
                continue
            elif "from_spreadsheet" in component and component["from_spreadsheet"]:
                results_array.append({
                    "point_num": point_num,
                    "exists": False,
                    "action": True,
                    "error": False,
                    "result": "Component found in spreadsheet, point will be imported"
                })
            else:
                results_array.append({
                    "point_num": point_num,
                    "exists": False,
                    "action": True,
                    "error": False,
                    "result": "Component exists, point will be imported"
                })

            point = self.get_point_from_parsed_row(row, component["id"], point_type)

            self.points_by_num[point_num] = point
            points_to_insert.append(point)

        if not self.dry:
            batch = []
            for x in points_to_insert:
                batch.append(x)
                if len(batch) == 200:
                    self.uow.component_points.insert_bulk(batch)
                    batch = []
            if len(batch) > 0:
                self.uow.component_points.insert_bulk(batch)

    def parse_spreadsheet_row(self, row, point_type):
        point_num = None

        try:
            point_num = unicode(int(row[1].value)).strip()
        except ValueError:
            if row[1].value[0:3] == point_type + "-":
                point_num = unicode(int(row[1].value[3:])).strip()
        except:
            return None

        if not point_num or len(point_num.strip()) == 0:
            return None
        point = {
            "component_num": unicode(row[0].value).strip(),
            "point_num": unicode(row[0].value).strip().replace(" ", "") + "-" + point_type + "-" + point_num,
            "point_code": unicode(row[2].value).strip(),
            "point_description": unicode(row[3].value).strip(),
            "point_id": unicode(uuid.uuid4())
        }
        if point_type in ["EP", "CP", "NP"]:
            point["point_units"] = unicode(row[4].value).strip()

        if point_type == "CP":
            point["point_formula"] = unicode(row[5].value).strip()

        return point


    def get_point_from_parsed_row(self, parsed_row, component_id, point_type):
        point = {
            "id": parsed_row["point_id"],
            "code": parsed_row["point_code"],
            "component_id": component_id,
            "component_num": parsed_row["component_num"],
            "component_point_num": parsed_row["point_num"],
            "description": parsed_row["point_description"],
            "point_type": point_type
        }

        if point_type == "EP":
            point["point_type_caption"] = "Energy Point"
            point["units"] = parsed_row["point_units"]
        elif point_type == "CP":
            point["formula"] = parsed_row["point_formula"]
            point["point_type_caption"] = "Calculated Point"
            point["units"] = parsed_row["point_units"]
        elif point_type == "PP":
            point["point_type_caption"] = "Position Point"
        elif point_type == "NP":
            point["point_type_caption"] = "Numeric Point"
            point["units"] = parsed_row["point_units"]
        elif point_type == "BP":
            point["point_type_caption"] = "Binary Point"

        return point

"""
class MasterSpreadsheetImporter:
    def __init__(self, fh):
        self.uow = UoW(None)
        self.wb = None
        self.fh = fh

        self.components_by_num = dict()
        self.points_by_num = dict()
        self.paragraphs_by_num = dict()

    def load_components_from_db(self):
        components = self.uow.components.get_all()
        self.components_by_num = dict((c["num"], c) for c in components)

    def load_points_from_db(self):
        points = self.uow.component_points.get_all()
        self.points_by_num = dict((p["component_point_num"], p) for p in points)

    def load_paragraphs_from_db(self):
        paragraphs = self.uow.paragraphs.get_all()
        self.paragraphs_by_num = dict((p["num"], p) for p in paragraphs)

    def import_components(self):
        ws = self.wb.get_sheet_by_name("Components")
        if ws is None:
            return

        rows = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            rows.append({
                "component_num": str(row[0].value).strip(),
                "description": unicode(row[1].value).strip(),
                "component_id": str(uuid.uuid4())
            })

        # by sorting the list, we ensure that parents will always
        # come before their children
        rows = sorted(rows, key=lambda x: x["component_num"])

        components_to_insert = []

        for row in rows:
            component_num = row["component_num"]
            description = row["description"]
            component_id = row["component_id"]

            if component_num in self.components_by_num:
                old_description = self.components_by_num[component_num]["description"]
                if description != old_description:
                    self.components_by_num[component_num]["description"] = description
                    components_to_insert.append(self.components_by_num[component_num])
                continue

            parent_num = component_num.strip()

            while len(parent_num) > 0 and parent_num[len(parent_num) - 1] not in [".", "-", " "]:
                parent_num = parent_num[0:len(parent_num)-1]

            parent_num = parent_num[0:len(parent_num)-1]

            if len(parent_num) == 5:
                # parent is level 2 which means it needs a level 2 multiple of 10
                parent_num = parent_num[0:4] + "0"

            try:
                parent = self.components_by_num[parent_num]
                parent_id = parent["id"]
            except:
                parent = None
                parent_id = None

            component = {
                "id": component_id,
                "description": description,
                "num": component_num,
                "mapping_child_ids": [],
                "mapping_parent_ids": [],
                "mapping_root": False,
                "protected": False,
                "structure_child_ids": [],
                "structure_parent_id": parent_id
            }
            components_to_insert.append(component)

            if parent is not None:
                parent["structure_child_ids"].append(component_id)

            self.components_by_num[component_num] = component

        batch = []
        for x in components_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.components.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.components.insert_bulk(batch)

    def insert_energy_points(self):
        ws = self.wb.get_sheet_by_name("Energy Points")
        if ws is None:
            return

        points_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            point_num = unicode(row[1].value).strip()
            point_code = unicode(row[2].value).strip()
            point_description = unicode(row[3].value).strip()
            point_units = unicode(row[4].value).strip()
            point_id = str(uuid.uuid4())

            point_num = component_num.replace(" ", "") + "-EP-" + point_num

            if component_num not in self.components_by_num:
                continue

            if point_num in self.points_by_num:
                continue

            component = self.components_by_num[component_num]

            point = {
                "id": point_id,
                "code": point_code,
                "component_id": component["id"],
                "component_num": component_num,
                "component_point_num": point_num,
                "description": point_description,
                "point_type": "EP",
                "point_type_caption": "Energy Point",
                "units": point_units
            }

            points_to_insert.append(point)

            self.points_by_num[point_num] = point

        batch = []
        for x in points_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.component_points.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.component_points.insert_bulk(batch)

    def insert_calculated_points(self):
        ws = self.wb.get_sheet_by_name("Calculated Points")
        if ws is None:
            return

        points_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            point_num = unicode(row[1].value).strip()
            point_code = unicode(row[2].value).strip()
            point_description = unicode(row[3].value).strip()
            point_units = unicode(row[4].value).strip()
            point_formula = unicode(row[5].value).strip()
            point_id = str(uuid.uuid4())

            point_num = component_num.replace(" ", "") + "-CP-" + point_num

            if component_num not in self.components_by_num:
                continue

            if point_num in self.points_by_num:
                continue

            component = self.components_by_num[component_num]

            point = {
                "id": point_id,
                "code": point_code,
                "component_id": component["id"],
                "component_num": component_num,
                "component_point_num": point_num,
                "description": point_description,
                "point_type": "CP",
                "point_type_caption": "Calculated Point",
                "units": point_units,
                "formula": point_formula
            }

            points_to_insert.append(point)

            self.points_by_num[point_num] = point

        batch = []
        for x in points_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.component_points.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.component_points.insert_bulk(batch)

    def insert_position_points(self):
        ws = self.wb.get_sheet_by_name("Position Points")
        if ws is None:
            return

        points_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            point_num = unicode(row[1].value).strip()
            point_code = unicode(row[2].value).strip()
            point_description = unicode(row[3].value).strip()
            point_id = str(uuid.uuid4())

            point_num = component_num.replace(" ", "") + "-PP-" + point_num

            if component_num not in self.components_by_num:
                continue

            if point_num in self.points_by_num:
                continue

            component = self.components_by_num[component_num]

            point = {
                "id": point_id,
                "code": point_code,
                "component_id": component["id"],
                "component_num": component_num,
                "component_point_num": point_num,
                "description": point_description,
                "point_type": "PP",
                "point_type_caption": "Position Point"
            }

            points_to_insert.append(point)

            self.points_by_num[point_num] = point

        batch = []
        for x in points_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.component_points.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.component_points.insert_bulk(batch)

    def insert_numeric_points(self):
        ws = self.wb.get_sheet_by_name("Numeric Points")
        if ws is None:
            return

        points_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            point_num = unicode(row[1].value).strip()
            point_code = unicode(row[2].value).strip()
            point_description = unicode(row[3].value).strip()
            point_units = unicode(row[4].value).strip()
            point_id = str(uuid.uuid4())

            point_num = component_num.replace(" ", "") + "-NP-" + point_num

            if component_num not in self.components_by_num:
                continue

            if point_num in self.points_by_num:
                continue

            component = self.components_by_num[component_num]

            point = {
                "id": point_id,
                "code": point_code,
                "component_id": component["id"],
                "component_num": component_num,
                "component_point_num": point_num,
                "description": point_description,
                "point_type": "NP",
                "point_type_caption": "Numeric Point",
                "units": point_units
            }

            points_to_insert.append(point)

            self.points_by_num[point_num] = point

        batch = []
        for x in points_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.component_points.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.component_points.insert_bulk(batch)

    def insert_binary_points(self):
        ws = self.wb.get_sheet_by_name("Binary Points")
        if ws is None:
            return

        points_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            point_num = unicode(row[1].value).strip()
            point_code = unicode(row[2].value).strip()
            point_description = unicode(row[3].value).strip()
            point_id = str(uuid.uuid4())

            point_num = component_num.replace(" ", "") + "-BP-" + point_num

            if component_num not in self.components_by_num:
                continue

            if point_num in self.points_by_num:
                continue

            component = self.components_by_num[component_num]

            point = {
                "id": point_id,
                "code": point_code,
                "component_id": component["id"],
                "component_num": component_num,
                "component_point_num": point_num,
                "description": point_description,
                "point_type": "BP",
                "point_type_caption": "Binary Point"
            }

            points_to_insert.append(point)

            self.points_by_num[point_num] = point

        batch = []
        for x in points_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.component_points.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.component_points.insert_bulk(batch)

    def insert_paragraphs(self, paragraph_type):
        if paragraph_type == "AR":
            ws = self.wb.get_sheet_by_name("Acceptance Requirements")
        elif paragraph_type == "CR":
            ws = self.wb.get_sheet_by_name("Commissioning Requirements")
        elif paragraph_type == "CS":
            ws = self.wb.get_sheet_by_name("Control Sequences")
        elif paragraph_type == "DR":
            ws = self.wb.get_sheet_by_name("Demand Response")
        elif paragraph_type == "FT":
            ws = self.wb.get_sheet_by_name("Functional Tests")
        elif paragraph_type == "LC":
            ws = self.wb.get_sheet_by_name("Load Curtailment")
        elif paragraph_type == "MR":
            ws = self.wb.get_sheet_by_name("Maintenance Requirements")
        elif paragraph_type == "PR":
            ws = self.wb.get_sheet_by_name("Project Requirements")
        elif paragraph_type == "RR":
            ws = self.wb.get_sheet_by_name("Roles and Responsibilities")
        else:
            return

        if ws is None:
            return

        paragraphs_to_insert = []

        first_row = True
        for row in ws.iter_rows():
            if first_row:
                first_row = False
                continue

            component_num = str(row[0].value).strip()
            paragraph_num = str(row[1].value).strip()
            paragraph_title = unicode(row[2].value).strip()
            paragraph_text = unicode(row[3].value).strip()
            paragraph_id = str(uuid.uuid4())

            paragraph_num = component_num.replace(" ", "") + "-" + paragraph_type + "-" + paragraph_num

            if component_num not in self.components_by_num:
                continue

            if paragraph_num in self.paragraphs_by_num:
                continue

            component = self.components_by_num[component_num]

            paragraph = {
                "id": paragraph_id,
                "component_id": component["id"],
                "description": paragraph_text,
                "num": paragraph_num,
                "title": paragraph_title,
                "type": paragraph_type
            }

            paragraphs_to_insert.append(paragraph)

            self.paragraphs_by_num[paragraph_num] = paragraph

        batch = []
        for x in paragraphs_to_insert:
            batch.append(x)
            if len(batch) == 200:
                self.uow.paragraphs.insert_bulk(batch)
                batch = []
        if len(batch) > 0:
            self.uow.paragraphs.insert_bulk(batch)




    def run(self):
        try:
            self.wb = load_workbook(filename=self.fh, use_iterators=True)
            self.load_components_from_db()
            self.import_components()

            #self.load_points_from_db()
            #self.insert_energy_points()
            #self.insert_calculated_points()
            #self.insert_position_points()
            #self.insert_numeric_points()
            #self.insert_binary_points()
            #self.load_paragraphs_from_db()
            #self.insert_paragraphs("AR")
            #self.insert_paragraphs("CR")
            #self.insert_paragraphs("CS")
            #self.insert_paragraphs("DR")
            #self.insert_paragraphs("FT")
            #self.insert_paragraphs("LC")
            #self.insert_paragraphs("MR")
            #self.insert_paragraphs("PR")
            #self.insert_paragraphs("RR")
        except:
            logging.exception("An error occurred importing from master spreadsheet")

if __name__ == "__main__":
    importer = MasterSpreadsheetImporter("C:\\Users\\sprice\\Documents\\Work\\Pathian\\Master component import template.xlsx")
    importer.run()

"""